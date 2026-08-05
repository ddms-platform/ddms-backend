# -*- coding: utf-8 -*-
"""
Fix ONLY information in DDMS_Report5_Unit_Test.xlsx:
- Do NOT modify styles/fonts/borders/merges/charts.
- Recompute summary counts on row 7 from actual matrix markers (P/F and N/A/B).
- Replace label strings using EN DASH (U+2013) with hyphen '-'.

Output: *_info_fixed.xlsx (keeps original untouched; makes a backup copy).
"""
from __future__ import annotations

import os
import shutil
from openpyxl import load_workbook


TGT = r"c:\Users\Hi\Downloads\DDMS_Report5_Unit_Test.xlsx"
OUT = r"c:\Users\Hi\Downloads\DDMS_Report5_Unit_Test_info_fixed.xlsx"
BACKUP = r"c:\Users\Hi\Downloads\DDMS_Report5_Unit_Test_backup_before_info_fixed.xlsx"

EN_DASH = "\u2013"  # '–'


def find_row_with_labels(ws, prefix: str, min_count: int = 2, col_start: int = 1, col_end: int = 50):
    """Return (row, col_min, col_max, count) where at that row cell value startswith prefix."""
    for r in range(1, min(ws.max_row, 300) + 1):
        cols = []
        for c in range(col_start, col_end + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith(prefix):
                cols.append(c)
        if len(cols) >= min_count:
            return r, min(cols), max(cols), len(cols)
    return None


def count_in_row(ws, row: int, col_min: int, col_max: int, target: str):
    return sum(1 for c in range(col_min, col_max + 1) if ws.cell(row, c).value == target)


def find_best_row_by_markers(ws, col_min: int, col_max: int, targets: list[str], search_start: int = 1):
    """Pick row with maximum (count of targets) within [col_min..col_max]."""
    best = None  # (row, count)
    for r in range(search_start, min(ws.max_row, 500) + 1):
        cnt = sum(1 for c in range(col_min, col_max + 1) if ws.cell(r, c).value in targets)
        if cnt > (best[1] if best else -1):
            best = (r, cnt)
    return best[0] if best else None


def fix_summary_for_sheet(ws):
    found_utc = find_row_with_labels(ws, "UTCID", min_count=2, col_start=1, col_end=40)
    if not found_utc:
        return False
    utc_row, col_min, col_max, utc_count = found_utc

    # Result row: where P/F appear most
    result_row = find_best_row_by_markers(ws, col_min, col_max, ["P", "F"], search_start=utc_row + 1)

    # Type row: where N/A/B appear most
    type_row = find_best_row_by_markers(ws, col_min, col_max, ["N", "A", "B"], search_start=utc_row + 1)

    if result_row is None or type_row is None:
        return False

    pass_count = count_in_row(ws, result_row, col_min, col_max, "P")
    fail_count = count_in_row(ws, result_row, col_min, col_max, "F")

    normal_count = count_in_row(ws, type_row, col_min, col_max, "N")
    abnormal_count = count_in_row(ws, type_row, col_min, col_max, "A")
    boundary_count = count_in_row(ws, type_row, col_min, col_max, "B")

    type_total = normal_count + abnormal_count + boundary_count
    untested_count = max(0, type_total - pass_count - fail_count)

    # Summary cells (row 7 are fixed in this workbook layout)
    ws["A7"].value = float(pass_count)
    ws["C7"].value = float(fail_count)
    ws["F7"].value = float(untested_count)
    ws["L7"].value = float(normal_count)
    ws["M7"].value = float(abnormal_count)
    ws["N7"].value = float(boundary_count)
    ws["O7"].value = float(type_total)

    return True


def main():
    if not os.path.exists(TGT):
        raise SystemExit(f"Missing target: {TGT}")
    if not os.path.exists(BACKUP):
        shutil.copyfile(TGT, BACKUP)
    wb = load_workbook(TGT)

    core_keep = {"Guideline", "Cover", "Functions", "Statistics"}
    feature_sheets = [s for s in wb.sheetnames if s not in core_keep]

    # 1) Fix EN DASH in labels
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 250), max_col=min(ws.max_column, 30)):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and EN_DASH in v:
                    cell.value = v.replace(EN_DASH, "-")

    # 2) Fix summary counts in row 7 for each feature sheet
    changed = 0
    for name in feature_sheets:
        ws = wb[name]
        ok = fix_summary_for_sheet(ws)
        if ok:
            changed += 1

    wb.save(OUT)
    print("Backup:", BACKUP)
    print("Saved:", OUT)
    print("Sheets fixed:", changed, "/", len(feature_sheets))


if __name__ == "__main__":
    main()

