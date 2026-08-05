# -*- coding: utf-8 -*-
"""
Sync ONLY values (and cell hyperlinks) from a "correct" Report5 workbook
into another workbook, without modifying formatting (fonts/borders/fills/merges/charts).

Usage: python _sync_report5_info_only.py
"""
from __future__ import annotations

import os
import shutil
from openpyxl import load_workbook


SRC = r"c:\Users\Hi\Downloads\Report5_Unit Test.xlsx"
TGT = r"c:\Users\Hi\Downloads\DDMS_Report5_Unit_Test.xlsx"


def sync_region(ws_src, ws_tgt, row_start, row_end, col_start, col_end, debug=False):
    updated = 0

    # Update values (do not set styles)
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            s_cell = ws_src.cell(r, c)
            t_cell = ws_tgt.cell(r, c)

            if s_cell.value != t_cell.value:
                t_cell.value = s_cell.value
                updated += 1

            # Hyperlinks are "info", not formatting. Copy if present in source.
            try:
                s_h = s_cell.hyperlink
                if s_h is not None and (t_cell.hyperlink is None or s_h.target != (t_cell.hyperlink.target if t_cell.hyperlink else None)):
                    t_cell.hyperlink = s_h
            except Exception:
                pass

    if debug:
        print(f"[{ws_tgt.title}] updated cells: {updated}")
    return updated


def sync_workbook(src_path: str, tgt_path: str):
    if not os.path.exists(src_path):
        raise SystemExit(f"Missing SRC: {src_path}")
    if not os.path.exists(tgt_path):
        raise SystemExit(f"Missing TGT: {tgt_path}")

    backup_path = tgt_path.replace(".xlsx", "_backup_before_sync.xlsx")
    if not os.path.exists(backup_path):
        shutil.copyfile(tgt_path, backup_path)

    wb_src = load_workbook(src_path)
    wb_tgt = load_workbook(tgt_path)

    # Sheet names to sync
    common = [s for s in wb_tgt.sheetnames if s in wb_src.sheetnames]
    print("Common sheets:", len(common))

    total_updated = 0
    for name in common:
        ws_src = wb_src[name]
        ws_tgt = wb_tgt[name]

        # Keep ranges reasonably bounded (faster + safer)
        if name == "Cover":
            total_updated += sync_region(ws_src, ws_tgt, 1, 120, 1, 10)
        elif name == "Functions":
            total_updated += sync_region(ws_src, ws_tgt, 1, 150, 1, 10)
        elif name == "Statistics":
            total_updated += sync_region(ws_src, ws_tgt, 1, 250, 1, 12)
        elif name.startswith("Function"):
            # Function matrices in template live roughly in A..O and rows up to ~80
            total_updated += sync_region(ws_src, ws_tgt, 1, 120, 1, 25)
        else:
            # For any other sheet, only sync small metadata region (avoid touching chart anchors)
            # You requested "only add info", so keep it minimal.
            total_updated += sync_region(ws_src, ws_tgt, 1, 30, 1, 12)

        # Do NOT touch merges/charts/layout/styles.

    out_path = tgt_path.replace(".xlsx", "_synced_info_only.xlsx")
    wb_tgt.save(out_path)
    print("Saved:", out_path)
    print("Done. (Cell value sync only)")


if __name__ == "__main__":
    sync_workbook(SRC, TGT)

